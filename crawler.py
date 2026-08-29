from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    NoSuchElementException, TimeoutException, StaleElementReferenceException,
    UnexpectedAlertPresentException
)
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta, date
import os
import time
import re
import random
import io
import base64
from urllib.parse import parse_qs, unquote

import requests
from bs4 import BeautifulSoup

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    import gspread
    from google.oauth2.service_account import Credentials as GoogleCredentials
except ImportError:
    gspread = None
    GoogleCredentials = None

# ========================= 식별정보 (전부 Secrets/환경변수에서 읽음) =========================
# 리포지토리가 Public이므로, 대상 사이트 도메인/물건 종류 명칭/개인 이메일/시트ID처럼
# 검색으로 찾아질 수 있는 값은 코드에 절대 리터럴로 적지 않는다. 전부 GitHub Secrets에
# 등록해두고 실행 시점에 환경변수로 읽어온다 - Secrets는 값이 로그에도 마스킹되고,
# 저장소 코드/커밋 히스토리 어디에도 실제 값이 남지 않는다.
SITE_DOMAIN = os.environ.get("SITE_DOMAIN", "")          # 예: 실제 사이트 도메인
SITE_BASE = f"https://{SITE_DOMAIN}" if SITE_DOMAIN else ""
TYPE_A_LABEL = os.environ.get("TYPE_A_LABEL", "TYPE_A")  # 예: 물건종류 A명칭
TYPE_B_LABEL = os.environ.get("TYPE_B_LABEL", "TYPE_B")  # 예: 물건종류 B명칭
TYPE_HEADER_LABEL = f"{TYPE_A_LABEL}/{TYPE_B_LABEL}"
LOGIN_ID = os.environ.get("SITE_LOGIN_ID", "")
LOGIN_PW = os.environ.get("SITE_LOGIN_PW", "")
GOOGLE_SHEET_ID = os.environ.get("GOOGLE_SHEET_ID", "")
GOOGLE_SHEET_SHARE_EMAIL = os.environ.get("SHARE_EMAIL", "")

# ========================= 사용자 설정 =========================
DAYS_AHEAD = 60
CTGR_VALUE = '30'
SAVE_DIR = "/tmp/"
MAX_PAGES = 500
LIST_PAGE_SIZE = "100"
TEMPLATE_HEADERS = [
    TYPE_HEADER_LABEL, '사건번호', '주소', '참고사항', '용도', '면적', '진행상태',
    '최저가율', '최저가', '감평가', '입찰일', '상세페이지', '신규여부',
    '지분 여부', '필지별 주소', 'pnu', '토지이용계획', '소유자', '키워드'
]

FETCH_DETAIL_DOCS = True
DETAIL_PAGE_DELAY = (1.0, 2.0)
PDF_DOWNLOAD_DELAY = (0.6, 1.5)
PAGE_CLICK_DELAY = (1.5, 3.0)
AC_COL = 29
CURRENCY_COLUMNS = (6, 9, 10)

GOOGLE_SHEETS_ENABLED = True
GOOGLE_CREDENTIALS_PATH_CANDIDATES = []
GOOGLE_SHEET_WEEK_RETENTION = 3
AUTOSAVE_EVERY_N_ITEMS = 50

# --- GitHub Actions 환경변수 연동 (수집 범위/감정가 상한) ---
COLLECT_MODE_ENV = os.environ.get("COLLECT_MODE", "both")
APSL_AMT_END_ENV = os.environ.get("APSL_AMT_END", "500000000")

# --- GitHub Actions 환경변수 연동 (수집 범위/감정가 상한/실행 트리거 종류) ---
COLLECT_MODE_ENV = os.environ.get("COLLECT_MODE", "both")
APSL_AMT_END_ENV = os.environ.get("APSL_AMT_END", "500000000")
# TRIGGER_TYPE: 워크플로우가 "schedule"(예약)로 시작됐는지, 그 외(수동 실행 등)로
# 시작됐는지를 나타낸다. 워크플로우 파일에서 github.event_name 값을 그대로 넘겨준다.
# 기본값을 "manual"로 둔 이유: 이 값을 못 받아오는 상황(예: 로컬에서 직접 실행,
# 워크플로우 설정 누락 등)에서는 안전한 쪽(=예약 전용 탭을 건드리지 않는 별도 탭)으로
# 동작하도록 하기 위함이다 - "schedule"로 잘못 오인해서 예약 데이터가 있는 탭을
# 건드리는 사고보다는, 별도 탭에 쓰는 게 항상 더 안전하다.
TRIGGER_TYPE_ENV = os.environ.get("TRIGGER_TYPE", "manual")

_creds_json_env = os.environ.get("GOOGLE_SA_KEY_JSON", "")
if _creds_json_env:
    _tmp_cred_path = "/tmp/google_credentials.json"
    with open(_tmp_cred_path, "w", encoding="utf-8") as f:
        f.write(_creds_json_env)
    GOOGLE_CREDENTIALS_PATH_CANDIDATES = [_tmp_cred_path] + GOOGLE_CREDENTIALS_PATH_CANDIDATES


def _resolve_google_credentials_path(candidates):
    for path in candidates:
        if os.path.exists(path):
            return path
    return candidates[0] if candidates else ""


GOOGLE_CREDENTIALS_PATH = _resolve_google_credentials_path(GOOGLE_CREDENTIALS_PATH_CANDIDATES)
# ================================================================

오늘 = datetime.now()
start_date_str = 오늘.strftime('%Y-%m-%d')
end_date_str = (오늘 + timedelta(days=DAYS_AHEAD)).strftime('%Y-%m-%d')
# 로컬 엑셀 파일명은 main()에서 실제 사용된 구글 시트 탭 이름(예약 전용 탭인지 수동
# 전용 탭인지에 따라 달라짐)에 맞춰 그 이름 그대로 ".xlsx"만 붙여서 결정한다 - 그래야
# 어떤 실행이 만든 파일인지 파일명만 보고도 바로 알 수 있고, 탭 이름과 파일명이 항상
# 정확히 짝을 이룬다.

BASE_URL = f"{SITE_BASE}/ca/caList.php?srchHistory=1"

PA_BASE_URL = f"{SITE_BASE}/pa/paList.php?srchHistory=1"
PA_CTGR_VALUE = '10100'
PA_PRPTDVSN_VALUES = [
    ('seizure', '압류재산'), ('other_general', '기타일반재산'),
    ('finance_collateral', '금융권담보재산'), ('bankruptcy', '파산재산'),
    ('public', '공유재산'), ('public_development', '공공개발재산'),
    ('acquired', '유입재산'), ('entrusted', '수탁재산'), ('disused', '불용품'),
]

AMOUNT_OPTIONS = [
    (1000000, "1백만"), (2000000, "2백만"), (3000000, "3백만"), (4000000, "4백만"),
    (5000000, "5백만"), (6000000, "6백만"), (7000000, "7백만"), (8000000, "8백만"),
    (9000000, "9백만"), (10000000, "1천만"), (20000000, "2천만"), (30000000, "3천만"),
    (40000000, "4천만"), (50000000, "5천만"), (60000000, "6천만"), (70000000, "7천만"),
    (80000000, "8천만"), (90000000, "9천만"), (100000000, "1억"), (150000000, "1억 5천만"),
    (200000000, "2억"), (250000000, "2억 5천만"), (300000000, "3억"), (350000000, "3억 5천만"),
    (400000000, "4억"), (450000000, "4억 5천만"), (500000000, "5억"), (600000000, "6억"),
    (700000000, "7억"), (800000000, "8억"), (900000000, "9억"), (1000000000, "10억"),
    (1100000000, "11억"), (1200000000, "12억"), (1300000000, "13억"), (1400000000, "14억"),
    (1500000000, "15억"), (1600000000, "16억"), (1700000000, "17억"), (1800000000, "18억"),
    (1900000000, "19억"), (2000000000, "20억"), (3000000000, "30억"), (4000000000, "40억"),
    (5000000000, "50억"), (6000000000, "60억"), (7000000000, "70억"), (8000000000, "80억"),
    (9000000000, "90억"), (10000000000, "100억"), (20000000000, "200억"), (30000000000, "300억"),
    (40000000000, "400억"), (50000000000, "500억"), (60000000000, "600억"), (70000000000, "700억"),
    (80000000000, "800억"), (90000000000, "900억"), (100000000000, "1000억"),
]
DEFAULT_APSL_AMT_END = 500000000


def append_row_with_format(sheet, row_values, ac_value=None):
    sheet.append(row_values)
    r = sheet.max_row
    for col in CURRENCY_COLUMNS:
        cell = sheet.cell(row=r, column=col)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.##'
    if ac_value:
        sheet.cell(row=r, column=AC_COL, value=ac_value)
    return r


def _iso_week_label(dt):
    iso_year, iso_week, _ = dt.isocalendar()
    return f"{iso_year}-W{iso_week:02d}"


def _week_tab_name(dt):
    """dt가 속한 ISO 주차의 "예약 실행 전용" 구글 시트 탭 이름을 만든다.
    형식은 "YYYYMMDD_W주차_data"이며, YYYYMMDD는 그 주의 **토요일** 날짜다
    (ISO 요일번호 6 = 토요일). 예약(schedule)으로 트리거된 실행만 이 탭을 쓴다."""
    iso_year, iso_week, _ = dt.isocalendar()
    saturday = date.fromisocalendar(iso_year, iso_week, 6)
    return f"{saturday.strftime('%Y%m%d')}_W{iso_week:02d}_data"


def _manual_tab_name(dt):
    """예약(schedule)이 아니라 사람이 직접(workflow_dispatch) 실행했을 때 쓰는 탭
    이름을 만든다: "YYYYMMDD_manual_data" (dt=실행일). 예약 실행 전용 탭
    (_week_tab_name)과 이름 패턴 자체가 달라서(주차 번호 "W.." 부분이 없음) 절대
    같은 탭을 가리키지 않는다 - 예약 실행 도중이든 끝난 뒤든 언제 수동 실행을 하더라도
    예약 데이터가 있는 탭을 절대 건드리지(=지우거나 덮어쓰지) 않는다. 같은 날 수동
    실행을 여러 번 하면(예: 실패해서 재시도) 그 날짜의 탭 하나에 자연스럽게 이어서
    쌓인다(증분 업데이트 로직이 그대로 적용됨)."""
    return f"{dt.strftime('%Y%m%d')}_manual_data"


WEEK_TAB_PATTERN = re.compile(r"^\d{8}_W\d{2}_data$")


def _read_service_account_email(credentials_path):
    try:
        import json as _json
        with open(credentials_path, "r", encoding="utf-8") as f:
            return _json.load(f).get("client_email", "")
    except Exception:
        return ""


def _validate_google_sheet_id(sheet_id):
    sid = (sheet_id or "").strip()
    if sid == "":
        return None, "GOOGLE_SHEET_ID가 비어있습니다."
    if sid.startswith("http"):
        return None, (f"GOOGLE_SHEET_ID에 URL 전체가 들어있습니다: {sid!r}")
    try:
        sid.encode("ascii")
    except UnicodeEncodeError:
        return None, f"GOOGLE_SHEET_ID에 한글/특수문자가 섞여 있습니다: {sid!r}"
    return sid, None


def get_google_spreadsheet():
    if not GOOGLE_SHEETS_ENABLED:
        return None
    if gspread is None or GoogleCredentials is None:
        print("[구글시트] gspread / google-auth 라이브러리가 설치되어 있지 않아 이번에는 건너뜁니다.")
        return None
    if not os.path.exists(GOOGLE_CREDENTIALS_PATH):
        print(f"[구글시트] 서비스 계정 키 파일을 찾을 수 없어 건너뜁니다: {GOOGLE_CREDENTIALS_PATH}")
        return None

    sheet_id, id_err = _validate_google_sheet_id(GOOGLE_SHEET_ID)
    if id_err and not GOOGLE_SHEET_ID:
        pass
    elif id_err:
        print(f"[구글시트] {id_err} 이번에는 건너뜁니다.")
        return None

    try:
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        gc = gspread.service_account(filename=GOOGLE_CREDENTIALS_PATH, scopes=scopes)
        service_account_email = _read_service_account_email(GOOGLE_CREDENTIALS_PATH)

        if sheet_id:
            try:
                sh = gc.open_by_key(sheet_id)
            except gspread.exceptions.APIError as e:
                print(f"[구글시트] 스프레드시트 접근 권한이 없어 이번에는 건너뜁니다 "
                      f"(GOOGLE_SHEET_ID={sheet_id!r}). 서비스 계정 이메일"
                      f"({service_account_email or '키 파일의 client_email 값'})을 '편집자'로 "
                      f"추가했는지 확인하세요. (원본 에러: {e})")
                return None
            except gspread.exceptions.SpreadsheetNotFound:
                print(f"[구글시트] GOOGLE_SHEET_ID로 스프레드시트를 찾을 수 없습니다: {sheet_id!r}")
                return None
        else:
            sh = gc.create("weekly_data_sheet")
            print(f"[구글시트] 새 스프레드시트를 생성했습니다: {sh.url}")
            if GOOGLE_SHEET_SHARE_EMAIL:
                try:
                    sh.share(GOOGLE_SHEET_SHARE_EMAIL, perm_type='user', role='writer')
                    print(f"[구글시트] {GOOGLE_SHEET_SHARE_EMAIL} 계정에 편집 권한을 공유했습니다.")
                except Exception as e:
                    print(f"[구글시트] 공유 실패: {e}")

        return sh
    except Exception as e:
        print(f"[구글시트] 인증/연결에 실패해 이번에는 건너뜁니다: {e}")
        return None


def get_or_create_tab(sh, tab_name):
    if sh is None:
        return None
    try:
        return sh.worksheet(tab_name)
    except gspread.exceptions.WorksheetNotFound:
        try:
            ws = sh.add_worksheet(title=tab_name, rows=100, cols=40)
            print(f"[구글시트] '{tab_name}' 탭이 없어 새로 만들었습니다.")
            return ws
        except Exception as e:
            print(f"[구글시트] '{tab_name}' 탭을 새로 만드는 데 실패했습니다: {e}")
            return None
    except Exception as e:
        print(f"[구글시트] '{tab_name}' 탭을 여는 데 실패했습니다: {e}")
        return None


def get_tab_if_exists(sh, tab_name):
    if sh is None:
        return None
    try:
        return sh.worksheet(tab_name)
    except gspread.exceptions.WorksheetNotFound:
        return None
    except Exception as e:
        print(f"[구글시트] '{tab_name}' 탭을 확인하는 데 실패했습니다: {e}")
        return None


def load_existing_data_from_gsheet(ws):
    if ws is None:
        return set(), []
    try:
        all_values = ws.get_all_values()
    except Exception as e:
        print(f"[구글시트] '{ws.title}' 탭을 읽는 데 실패했습니다: {e} - 중복 검사 없이 진행합니다.")
        return set(), []

    if not all_values:
        return set(), []

    header_row = all_values[0]
    header_idx = {name: i for i, name in enumerate(header_row) if name}
    if "사건번호" not in header_idx:
        print(f"[구글시트] '{ws.title}' 탭에서 '사건번호' 열을 찾지 못해 중복 검사 없이 진행합니다.")
        return set(), []

    ac_idx = None
    for i, name in enumerate(header_row):
        if name and "토지등기" in str(name):
            ac_idx = i
            break

    currency_positions = {c - 1 for c in CURRENCY_COLUMNS}
    sano_i = header_idx["사건번호"]
    existing_case_numbers = set()
    existing_rows = []
    for row in all_values[1:]:
        if sano_i >= len(row):
            continue
        사건번호 = row[sano_i]
        if not 사건번호:
            continue
        existing_case_numbers.add(str(사건번호).strip())
        row_values = [
            row[header_idx[h]] if (h in header_idx and header_idx[h] < len(row)) else ""
            for h in TEMPLATE_HEADERS
        ]
        for i in currency_positions:
            if i < len(row_values) and row_values[i] not in ("", None):
                try:
                    cleaned = str(row_values[i]).replace(",", "").strip()
                    row_values[i] = float(cleaned) if "." in cleaned else int(cleaned)
                except (ValueError, TypeError):
                    pass
        ac_value = row[ac_idx] if (ac_idx is not None and ac_idx < len(row)) else None
        existing_rows.append((row_values, ac_value))

    return existing_case_numbers, existing_rows


def load_case_numbers_by_type_from_gsheet(ws):
    if ws is None:
        return set(), set()
    try:
        all_values = ws.get_all_values()
    except Exception:
        return set(), set()
    if not all_values:
        return set(), set()

    header_row = all_values[0]
    header_idx = {name: i for i, name in enumerate(header_row) if name}
    if TYPE_HEADER_LABEL not in header_idx or "사건번호" not in header_idx:
        return set(), set()

    jongryu_i = header_idx[TYPE_HEADER_LABEL]
    sano_i = header_idx["사건번호"]
    ca_set, pa_set = set(), set()
    for row in all_values[1:]:
        if sano_i >= len(row) or jongryu_i >= len(row):
            continue
        사건번호 = (row[sano_i] or "").strip()
        if not 사건번호:
            continue
        종류 = row[jongryu_i]
        if 종류 == TYPE_A_LABEL:
            ca_set.add(사건번호)
        elif 종류 == TYPE_B_LABEL:
            pa_set.add(사건번호)
    return ca_set, pa_set


def cleanup_old_week_tabs(sh, retention=GOOGLE_SHEET_WEEK_RETENTION):
    if sh is None:
        return 0
    try:
        titles = [ws.title for ws in sh.worksheets()]
    except Exception as e:
        print(f"[구글시트] 탭 목록을 가져오는 데 실패해 오래된 탭 정리를 건너뜁니다: {e}")
        return 0

    week_titles = sorted({t for t in titles if WEEK_TAB_PATTERN.match(t)}, reverse=True)
    to_delete = [t for t in week_titles if t not in week_titles[:retention]]
    if not to_delete:
        return 0

    removed = 0
    for title in to_delete:
        try:
            ws = sh.worksheet(title)
            sh.del_worksheet(ws)
            removed += 1
            print(f"[구글시트] 보관 기간({retention}주)이 지난 오래된 주차 탭을 삭제했습니다: {title}")
        except Exception as e:
            print(f"[구글시트] '{title}' 탭 삭제에 실패했습니다: {e}")
    return removed


def _gsheet_json_safe(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return value


def _apply_gsheet_formatting(ws, n_rows, n_cols):
    _apply_gsheet_formatting_range(ws, 1, n_rows, n_cols)


def _apply_gsheet_formatting_range(ws, start_row, end_row, n_cols):
    try:
        last_col = get_column_letter(n_cols)
        ws.format(f"A{start_row}:{last_col}{end_row}", {"wrapStrategy": "CLIP"})
        currency_start = max(start_row, 2)
        if end_row >= currency_start:
            for col in CURRENCY_COLUMNS:
                if col > n_cols:
                    continue
                col_letter = get_column_letter(col)
                ws.format(f"{col_letter}{currency_start}:{col_letter}{end_row}",
                          {"numberFormat": {"type": "NUMBER", "pattern": "#,##0.##"}})
    except Exception as e:
        print(f"[구글시트] 서식(줄바꿈/콤마) 적용 중 오류가 발생했지만 데이터 저장에는 영향이 없습니다: {e}")


def sync_to_google_sheet(local_sheet, ws):
    if ws is None:
        return False
    try:
        values = []
        for row in local_sheet.iter_rows(values_only=True):
            values.append([_gsheet_json_safe(v) for v in row])

        n_rows = max(len(values), 1)
        n_cols = max((len(r) for r in values), default=1)
        n_cols = max(n_cols, AC_COL)
        needed_rows = n_rows + 10
        needed_cols = n_cols + 2
        if ws.row_count < needed_rows or ws.col_count < needed_cols:
            ws.resize(rows=max(ws.row_count, needed_rows), cols=max(ws.col_count, needed_cols))

        ws.clear()
        ws.update(values=values, range_name="A1")
        _apply_gsheet_formatting(ws, n_rows, n_cols)
        print(f"[구글시트] 동기화 완료: {len(values) - 1 if values else 0}건 "
              f"(스프레드시트 ID: {GOOGLE_SHEET_ID or '(신규 생성)'} / 탭: {ws.title})")
        return True
    except Exception as e:
        print(f"[구글시트] 업로드 중 오류가 발생해 이번에는 건너뜁니다: {e}")
        return False


def append_new_rows_to_gsheet(local_sheet, ws, sync_state):
    if ws is None:
        return False
    try:
        last_synced = sync_state.get("synced_rows", 1)
        max_row = local_sheet.max_row
        if max_row <= last_synced:
            return True

        new_values = []
        for row in local_sheet.iter_rows(min_row=last_synced + 1, max_row=max_row, values_only=True):
            new_values.append([_gsheet_json_safe(v) for v in row])

        n_cols = max((len(r) for r in new_values), default=1)
        n_cols = max(n_cols, AC_COL)
        needed_rows = max_row + 10
        needed_cols = n_cols + 2
        if ws.row_count < needed_rows or ws.col_count < needed_cols:
            ws.resize(rows=max(ws.row_count, needed_rows), cols=max(ws.col_count, needed_cols))

        ws.append_rows(new_values, value_input_option="RAW")
        _apply_gsheet_formatting_range(ws, last_synced + 1, max_row, n_cols)
        sync_state["synced_rows"] = max_row
        print(f"[구글시트] 증분 동기화 완료: {len(new_values)}건 추가 (누적 {max_row - 1}건, 탭: {ws.title})")
        return True
    except Exception as e:
        print(f"[구글시트] 증분 업로드 중 오류가 발생해 이번에는 건너뜁니다: {e}")
        return False


def maybe_autosave(sheet, ws, save_path, total_new_count, sync_state):
    if total_new_count <= 0 or total_new_count % AUTOSAVE_EVERY_N_ITEMS != 0:
        return
    print(f"[자동 저장] 지금까지 새로 수집한 물건 {total_new_count}건 - 중간 저장 중...")
    if GOOGLE_SHEETS_ENABLED and ws is not None:
        append_new_rows_to_gsheet(sheet, ws, sync_state)
    try:
        sheet.parent.save(save_path)
        print(f"[자동 저장] 로컬 엑셀에도 중간 저장했습니다: {save_path}")
    except Exception as e:
        print(f"[자동 저장] 로컬 엑셀 중간 저장에 실패했습니다: {e}")


def prompt_collect_mode():
    """GitHub Actions 환경변수 COLLECT_MODE로 대체 (type_a / type_b / both)"""
    mode = COLLECT_MODE_ENV
    if mode not in ("type_a", "type_b", "both"):
        print(f"알 수 없는 COLLECT_MODE 값({mode!r}) - 기본값(both)으로 진행합니다.")
        return "both"
    print(f"[설정] 수집 범위: {mode} (환경변수 COLLECT_MODE)")
    return mode


def prompt_apsl_amt_end():
    """GitHub Actions 환경변수 APSL_AMT_END(원 단위)로 대체"""
    label_map = {label: val for val, label in AMOUNT_OPTIONS}
    valid_values = {val for val, _ in AMOUNT_OPTIONS}
    raw = APSL_AMT_END_ENV.strip()
    if raw in label_map:
        result = str(label_map[raw])
    elif raw.isdigit() and int(raw) in valid_values:
        result = raw
    else:
        print(f"[설정] APSL_AMT_END 값({raw!r})이 올바르지 않아 기본값(5억)으로 진행합니다.")
        result = str(DEFAULT_APSL_AMT_END)
    print(f"[설정] 감정가격 상한: {int(result):,}원 (환경변수 APSL_AMT_END)")
    return result


def find_chrome_binary():
    candidates = [
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
        "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
        "/usr/bin/google-chrome",
        "/usr/bin/google-chrome-stable",
        "/usr/bin/chromium-browser",
        "/usr/bin/chromium",
    ]
    for path in candidates:
        if path and os.path.exists(path):
            return path
    return None


def get_driver():
    chrome_options = Options()
    chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])
    chrome_options.add_argument("--window-size=1920,1080")
    # GitHub Actions(리눅스 러너, 화면 없음)에서 반드시 headless로 실행
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")

    chrome_path = find_chrome_binary()
    if chrome_path:
        chrome_options.binary_location = chrome_path
        print(f"구글 크롬 실행 파일 사용: {chrome_path}")
    else:
        print("구글 크롬 실행 파일을 자동으로 찾지 못해 기본 설정으로 진행합니다.")

    service = Service()
    return webdriver.Chrome(service=service, options=chrome_options)


def dismiss_alert_if_present(driver, tid=None):
    try:
        alert = driver.switch_to.alert
        text = alert.text
        alert.accept()
        print(f"  [tid={tid}] 알림창 발견, 닫음: {text}")
        return text
    except Exception:
        return None


def login(driver, login_id, password):
    if not login_id or not password:
        print("로그인 정보가 입력되지 않아 로그인을 건너뜁니다 (상세페이지 관련 항목은 비어있게 됩니다).")
        return False
    try:
        driver.execute_script("""
            const btn = document.querySelector('span.topGrid[data-action="loginDivBtn"]');
            if (btn) btn.click();
        """)
        WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.ID, "client_id"))
        )
        id_el = driver.find_element(By.ID, "client_id")
        pw_el = driver.find_element(By.ID, "passwd")
        id_el.clear()
        id_el.send_keys(login_id)
        pw_el.clear()
        pw_el.send_keys(password)
        driver.find_element(By.ID, "loginBtn").click()
        time.sleep(1.5)
        dismiss_alert_if_present(driver)
        try:
            login_btn_still = driver.find_elements(
                By.CSS_SELECTOR, 'span.topGrid[data-action="loginDivBtn"]'
            )
            if login_btn_still and login_btn_still[0].is_displayed():
                print("경고: 로그인이 안 된 것 같습니다 (아이디/비밀번호를 확인해주세요).")
                return False
        except Exception:
            pass
        print("로그인 완료(또는 시도됨) - 이어서 진행합니다.")
        return True
    except TimeoutException:
        print("경고: 로그인 창(아이디 입력란)을 찾지 못했습니다.")
        return False
    except Exception as e:
        print(f"로그인 시도 중 오류: {e}")
        return False


def set_field_value(driver, selector, value, label):
    try:
        existed = driver.execute_script("""
            const el = document.querySelector(arguments[0]);
            if (!el) return false;
            el.value = arguments[1];
            el.dispatchEvent(new Event('input', {bubbles: true}));
            el.dispatchEvent(new Event('change', {bubbles: true}));
            return true;
        """, selector, value)
        if existed:
            print(f"{label} 설정 완료: {value}")
        else:
            print(f"{label} 설정 실패: {selector} 요소를 페이지에서 찾을 수 없습니다.")
    except Exception as e:
        print(f"{label} 설정 중 오류 발생: {e}")


def set_radio_value(driver, name, value, label):
    try:
        existed = driver.execute_script("""
            const el = document.querySelector(
                'input[name="' + arguments[0] + '"][value="' + arguments[1] + '"]'
            );
            if (!el) return false;
            el.checked = true;
            el.dispatchEvent(new Event('click', {bubbles: true}));
            el.dispatchEvent(new Event('change', {bubbles: true}));
            return true;
        """, name, value)
        if existed:
            print(f"{label} 설정 완료: {value}")
        else:
            print(f"{label} 설정 실패: input[name={name}][value={value}] 요소를 찾을 수 없습니다.")
    except Exception as e:
        print(f"{label} 설정 중 오류 발생: {e}")


def set_checkbox_value(driver, selector, checked, label):
    try:
        existed = driver.execute_script("""
            const el = document.querySelector(arguments[0]);
            if (!el) return false;
            el.checked = arguments[1];
            el.dispatchEvent(new Event('click', {bubbles: true}));
            el.dispatchEvent(new Event('change', {bubbles: true}));
            return true;
        """, selector, checked)
        if existed:
            print(f"{label} 설정 완료: {'체크' if checked else '해제'}")
        else:
            print(f"{label} 설정 실패: {selector} 요소를 페이지에서 찾을 수 없습니다.")
    except Exception as e:
        print(f"{label} 설정 중 오류 발생: {e}")


def set_page_size(driver, size=LIST_PAGE_SIZE):
    set_field_value(driver, "#dataSize", str(size), f"목록 표시 개수({size}개씩 보기)")
    time.sleep(1.5)


def apply_filters(driver, apsl_amt_end_value):
    set_field_value(driver, "#ctgr", CTGR_VALUE, "물건종류")
    time.sleep(1.2)
    set_field_value(driver, "#bgnDt", start_date_str, "매각기일 시작일")
    time.sleep(1.2)
    set_field_value(driver, "#endDt", end_date_str, "매각기일 종료일")
    time.sleep(1.2)
    set_field_value(driver, "#apslAmtEnd", apsl_amt_end_value, "감정가격 상한")
    time.sleep(2)


MULTI_PARCEL_PATTERN = re.compile(r'외\s*(\d+)\s*필지')


def is_multi_parcel(address: str) -> bool:
    return bool(MULTI_PARCEL_PATTERN.search(address or ""))


def collect_share_tids(driver, apsl_amt_end_value):
    apply_filters(driver, apsl_amt_end_value)
    wait_for_list(driver)
    set_checkbox_value(driver, '#chkSpl_121', True, "[사전조회] 지분입찰 물건 필터")
    time.sleep(1.5)
    wait_for_list(driver)
    set_page_size(driver)
    wait_for_list(driver)

    share_tids = set()
    page_num = 1
    while page_num <= MAX_PAGES:
        wait_for_list(driver)
        time.sleep(0.3)
        rows = get_rows(driver)
        if not rows:
            break
        prev_first_tid = rows[0].get_attribute("data-tid")
        for row in rows:
            tid = row.get_attribute("data-tid")
            if tid:
                share_tids.add(tid)
        next_page = page_num + 1
        if not go_to_page(driver, next_page, prev_first_tid):
            break
        page_num = next_page

    set_checkbox_value(driver, '#chkSpl_121', False, "[사전조회] 지분입찰 물건 필터")
    time.sleep(1.0)
    print(f"[사전조회] 지분 물건 Tid {len(share_tids)}개 확인 (팝업 방문 없이 목록만으로 수집)")
    return share_tids


def apply_filters_pa(driver, apsl_amt_end_value, prptdvsn_value):
    set_field_value(driver, "#ctgr", PA_CTGR_VALUE, f"[{TYPE_B_LABEL}] 물건종류")
    time.sleep(1.0)
    set_field_value(driver, "#bgnDt", start_date_str, f"[{TYPE_B_LABEL}] 입찰일자 시작일")
    time.sleep(1.0)
    set_field_value(driver, "#clsDt", end_date_str, f"[{TYPE_B_LABEL}] 입찰일자 종료일")
    time.sleep(1.0)
    set_radio_value(driver, "dpslCd", "1", f"[{TYPE_B_LABEL}] 처분방식(매각)")
    time.sleep(1.0)
    set_field_value(driver, "#prptDvsn", prptdvsn_value, f"[{TYPE_B_LABEL}] 자산구분")
    time.sleep(1.0)
    set_field_value(driver, "#apslAmtEnd", str(apsl_amt_end_value), f"[{TYPE_B_LABEL}] 감정가격 상한")
    time.sleep(2.0)


def wait_for_list(driver, timeout=15):
    WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.ID, "lsTbody"))
    )


def get_rows(driver):
    return driver.find_elements(By.CSS_SELECTOR, "#lsTbody > tr[data-tid]")


def get_first_tid(driver):
    rows = get_rows(driver)
    return rows[0].get_attribute("data-tid") if rows else None


def extract_row(row, tid, chk_no=None, tot_no=None):
    try:
        try:
            addr_el = row.find_element(By.CSS_SELECTOR, f"#adrs_{tid}")
            주소 = (addr_el.get_attribute("textContent") or "").strip()
        except NoSuchElementException:
            주소 = ""

        try:
            sano_el = row.find_element(By.CSS_SELECTOR, f"#saNo_{tid}")
            raw = sano_el.get_attribute("textContent") or ""
            사건번호 = re.sub(r"\s+", " ", raw.split("지도")[0]).strip()
        except NoSuchElementException:
            사건번호 = ""

        try:
            ctgr_el = row.find_element(By.CSS_SELECTOR, f"#ctgr_{tid}")
            용도 = (ctgr_el.get_attribute("textContent") or "").strip()
        except NoSuchElementException:
            용도 = ""

        면적 = None
        try:
            area_spans = row.find_elements(By.CSS_SELECTOR, "div.gray.f12 span.gray")
            area_text = " ".join((s.get_attribute("textContent") or "") for s in area_spans)
            m = re.search(r"토지\s*([\d,\.]+)\s*㎡", area_text) or re.search(r"건물\s*([\d,\.]+)\s*㎡", area_text)
            if m:
                면적 = float(m.group(1).replace(",", ""))
        except Exception:
            pass

        참고사항 = "-"
        try:
            note_els = row.find_elements(By.CSS_SELECTOR, "div.f12 span.orange.f12")
            note_text = (note_els[0].get_attribute("textContent") or "").strip() if note_els else ""
            if note_text:
                참고사항 = note_text
        except Exception:
            pass

        감평가 = None
        최저가 = None
        try:
            apsl_el = row.find_element(By.CSS_SELECTOR, f"#apslAmt_{tid}")
            v = (apsl_el.text or "").replace(",", "").strip()
            감평가 = int(v) if v.isdigit() else None
        except NoSuchElementException:
            pass
        try:
            minb_el = row.find_element(By.CSS_SELECTOR, f"#minbAmt_{tid}")
            v = (minb_el.text or "").replace(",", "").strip()
            최저가 = int(v) if v.isdigit() else None
        except NoSuchElementException:
            pass

        진행상태 = ""
        유찰횟수 = 0
        try:
            stat_el = row.find_element(By.CSS_SELECTOR, f"td.center.no-950 #statNm_{tid}")
            진행상태 = (stat_el.get_attribute("textContent") or stat_el.text or "").strip()
            m2 = re.search(r"유찰\s*(\d+)\s*회", 진행상태)
            유찰횟수 = int(m2.group(1)) if m2 else 0
        except NoSuchElementException:
            pass
        최저가율 = None
        try:
            rate_el = row.find_element(By.CSS_SELECTOR, f"td.center.no-950 #statNm_{tid} + div")
            rate_text = (rate_el.get_attribute("textContent") or rate_el.text or "")
            rate_match = re.search(r"\d+", rate_text)
            최저가율 = int(rate_match.group(0)) if rate_match else None
        except NoSuchElementException:
            pass

        입찰일 = None
        try:
            date_el = row.find_element(By.CSS_SELECTOR, f"td.center.no-950 #bidDt_{tid}")
            raw_date = date_el.text.strip()
            time_el = row.find_element(By.CSS_SELECTOR, f"td.center.no-950 #bidDt_{tid} + div")
            raw_time = time_el.text.strip().strip("()")
            yy, mm, dd = raw_date.split(".")
            hh, mi = raw_time.split(":")
            입찰일 = datetime(2000 + int(yy), int(mm), int(dd), int(hh), int(mi))
        except Exception:
            pass

        신규여부 = "신규물건" if 유찰횟수 == 0 else "기존물건"
        지분여부 = "지분" if "지분" in 참고사항 else ""

        상세페이지 = ""
        if chk_no is not None and tot_no is not None:
            상세페이지 = f"{SITE_BASE}/ca/caView.php?tid={tid}&chkNo={chk_no}&TotNo={tot_no}"

        필지별주소 = ""
        pnu = ""
        토지이용계획 = ""
        소유자 = ""
        키워드 = ""

        return [
            TYPE_A_LABEL, 사건번호, 주소, 참고사항, 용도, 면적, 진행상태, 최저가율,
            최저가, 감평가, 입찰일, 상세페이지, 신규여부, 지분여부,
            필지별주소, pnu, 토지이용계획, 소유자, 키워드,
        ]
    except Exception as e:
        print(f"[tid={tid}] 데이터 추출 중 오류 발생: {e}")
        return None


def resolve_pa_year(mm, dd, reference_date):
    year = reference_date.year
    try:
        candidate = datetime(year, mm, dd)
    except ValueError:
        return None
    if candidate < reference_date - timedelta(days=30):
        candidate = datetime(year + 1, mm, dd)
    return candidate


def extract_row_pa(row, tid, chk_no=None, tot_no=None):
    try:
        try:
            mgmt_el = row.find_element(By.CSS_SELECTOR, f"#mgmtNo_{tid}")
            raw = mgmt_el.get_attribute("textContent") or ""
            m0 = re.search(r"\d{4}-\d+-\d+", raw)
            사건번호 = m0.group(0) if m0 else re.sub(r"\s+", " ", raw).strip()
        except NoSuchElementException:
            사건번호 = ""

        try:
            addr_el = row.find_element(By.CSS_SELECTOR, f"#cltrNm_{tid}")
            주소 = (addr_el.get_attribute("textContent") or "").strip()
        except NoSuchElementException:
            주소 = ""

        try:
            ctgr_el = row.find_element(By.CSS_SELECTOR, f"#ctgr_{tid}")
            용도 = (ctgr_el.get_attribute("textContent") or "").strip()
        except NoSuchElementException:
            용도 = ""

        면적 = None
        try:
            area_el = row.find_element(By.CSS_SELECTOR, f"#area_{tid}")
            area_text = area_el.get_attribute("textContent") or ""
            m1 = re.search(r"토지\s*([\d,\.]+)\s*㎡", area_text)
            if m1:
                면적 = float(m1.group(1).replace(",", ""))
        except NoSuchElementException:
            pass

        참고사항 = "-"
        try:
            note_els = row.find_elements(By.CSS_SELECTOR, "span.orange.f12")
            note_text = (note_els[0].get_attribute("textContent") or "").strip() if note_els else ""
            if note_text:
                참고사항 = note_text
        except Exception:
            pass

        감평가 = None
        try:
            apsl_el = row.find_element(By.CSS_SELECTOR, f"td.right.bold.no-950 #apslAmt_{tid}")
            v = re.sub(r"[^\d]", "", apsl_el.get_attribute("textContent") or "")
            감평가 = int(v) if v else None
        except NoSuchElementException:
            pass
        최저가 = None
        try:
            minb_el = row.find_element(By.CSS_SELECTOR, f"td.right.bold.no-950 #minbAmt_{tid}")
            v = re.sub(r"[^\d]", "", minb_el.get_attribute("textContent") or "")
            최저가 = int(v) if v else None
        except NoSuchElementException:
            pass

        진행상태 = ""
        유찰횟수 = 0
        try:
            stat_el = row.find_element(By.CSS_SELECTOR, f"td.center.no-950 #statNm_{tid}")
            진행상태 = (stat_el.get_attribute("textContent") or "").strip()
            m2 = re.search(r"유찰\s*(\d+)\s*회", 진행상태)
            유찰횟수 = int(m2.group(1)) if m2 else 0
        except NoSuchElementException:
            pass
        최저가율 = None
        try:
            rate_el = row.find_element(By.CSS_SELECTOR, f"td.center.no-950 #statNm_{tid} ~ span.blue")
            rate_text = rate_el.get_attribute("textContent") or ""
            rate_match = re.search(r"\d+", rate_text)
            최저가율 = int(rate_match.group(0)) if rate_match else None
        except NoSuchElementException:
            pass

        입찰일 = None
        try:
            date_el = row.find_element(By.CSS_SELECTOR, f"td.center.no-950 #bgnDtm_{tid}")
            raw_dt = (date_el.get_attribute("textContent") or "").strip()
            m3 = re.match(r"(\d{2})\.(\d{2})\s+(\d{2}):(\d{2})", raw_dt)
            if m3:
                mm, dd, hh, mi = (int(x) for x in m3.groups())
                candidate = resolve_pa_year(mm, dd, 오늘)
                if candidate:
                    입찰일 = candidate.replace(hour=hh, minute=mi)
        except Exception:
            pass

        신규여부 = "신규물건" if 유찰횟수 == 0 else "기존물건"
        지분여부 = "지분" if "지분" in 참고사항 else ""

        상세페이지 = ""
        if chk_no is not None and tot_no is not None:
            상세페이지 = f"{SITE_BASE}/pa/paView.php?cltrNo={tid}&chkNo={chk_no}&TotNo={tot_no}"

        필지별주소 = ""
        pnu = ""
        토지이용계획 = ""
        소유자 = ""
        키워드 = ""

        return [
            TYPE_B_LABEL, 사건번호, 주소, 참고사항, 용도, 면적, 진행상태, 최저가율,
            최저가, 감평가, 입찰일, 상세페이지, 신규여부, 지분여부,
            필지별주소, pnu, 토지이용계획, 소유자, 키워드,
        ]
    except Exception as e:
        print(f"[{TYPE_B_LABEL}][tid={tid}] 데이터 추출 중 오류 발생: {e}")
        return None


def build_requests_session(driver):
    session = requests.Session()
    try:
        ua = driver.execute_script("return navigator.userAgent;")
    except Exception:
        ua = "Mozilla/5.0"
    session.headers.update({"User-Agent": ua, "Referer": BASE_URL})
    sync_session_cookies(driver, session)
    return session


def sync_session_cookies(driver, session):
    for c in driver.get_cookies():
        try:
            session.cookies.set(c["name"], c["value"], domain=c.get("domain"))
        except Exception:
            pass


def open_detail_tab(driver, detail_url, tid=None):
    original_handle = driver.current_window_handle
    driver.execute_script("window.open(arguments[0], '_blank');", detail_url)
    try:
        WebDriverWait(driver, 10).until(lambda d: len(d.window_handles) > 1)
    except UnexpectedAlertPresentException:
        text = dismiss_alert_if_present(driver, tid)
        print(f"  [tid={tid}] 새 탭이 열리기 전에 알림창 발생({text}) - 로그인 상태를 확인해주세요.")
        return None, ""
    except Exception as e:
        print(f"  [tid={tid}] 상세페이지 탭을 여는 중 오류: {e}")
        return None, ""

    new_handle = [h for h in driver.window_handles if h != original_handle][-1]
    driver.switch_to.window(new_handle)
    html = ""
    try:
        WebDriverWait(driver, 15).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        time.sleep(0.8)
        html = driver.page_source
    except UnexpectedAlertPresentException:
        text = dismiss_alert_if_present(driver, tid)
        print(f"  [tid={tid}] 상세페이지 로딩 중 알림창 발생({text}) - 로그인 상태를 확인해주세요.")
    except Exception as e:
        print(f"  [tid={tid}] 상세페이지 로드 중 오류: {e}")
    return new_handle, html


def close_detail_tab(driver, handle, original_handle, tid=None):
    dismiss_alert_if_present(driver, tid)
    try:
        if handle and handle in driver.window_handles:
            driver.switch_to.window(handle)
            driver.close()
    except Exception:
        pass
    try:
        if original_handle in driver.window_handles:
            driver.switch_to.window(original_handle)
        elif driver.window_handles:
            driver.switch_to.window(driver.window_handles[0])
    except Exception:
        pass


def build_pdf_viewer_url(data_param):
    if not data_param:
        return None
    path = data_param if data_param.startswith("/") else "/" + data_param
    return SITE_BASE + path


def extract_raw_pdf_url(data_param):
    if not data_param:
        return None
    m = re.search(r"[?&]file=([^&]+)", data_param)
    if not m:
        return None
    file_path = m.group(1)
    if not file_path.startswith("/"):
        file_path = "/" + file_path
    return SITE_BASE + file_path


def click_registry_pdf_link(driver, detail_tab_handle, tid=None, timeout=10):
    try:
        link = driver.find_element(By.CSS_SELECTOR, 'a.btn-fileView[data-ctgr="DA"]')
    except NoSuchElementException:
        print(f"  [tid={tid}] 상세페이지에서 토지등기(DA) 링크를 찾지 못했습니다.")
        return None

    before_handles = set(driver.window_handles)
    try:
        driver.execute_script("arguments[0].click();", link)
    except UnexpectedAlertPresentException:
        text = dismiss_alert_if_present(driver, tid)
        print(f"  [tid={tid}] 토지등기 링크 클릭 중 알림창 발생({text})")
        return None
    except Exception as e:
        print(f"  [tid={tid}] 토지등기 링크 클릭 중 오류: {e}")
        return None

    pdf_url = None
    new_handle = None
    try:
        WebDriverWait(driver, timeout).until(lambda d: len(d.window_handles) > len(before_handles))
        new_handle = [h for h in driver.window_handles if h not in before_handles][-1]
        driver.switch_to.window(new_handle)
        WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        time.sleep(1.0)

        viewer_html = driver.page_source
        soup = BeautifulSoup(viewer_html, "html.parser")
        da_el = soup.select_one('div.file_menu.fileView[data-ctgr="DA"]')
        if da_el and da_el.get("data-param"):
            data_param = da_el["data-param"]
            pdf_url = extract_raw_pdf_url(data_param)
            if pdf_url:
                print(f"  [tid={tid}] 토지등기 팝업에서 실제 PDF 원본 URL 확인: {pdf_url}")
            else:
                pdf_url = build_pdf_viewer_url(data_param)
                print(f"  [tid={tid}] data-param에서 file= 파라미터를 못 찾아 뷰어 URL을 대신 사용: {pdf_url}")
        else:
            pdf_url = driver.current_url
            print(f"  [tid={tid}] 토지등기 팝업에서 data-param을 못 찾아 탭 URL을 그대로 사용: {pdf_url}")
    except UnexpectedAlertPresentException:
        text = dismiss_alert_if_present(driver, tid)
        print(f"  [tid={tid}] 토지등기 PDF 탭 로딩 중 알림창 발생({text})")
    except Exception as e:
        print(f"  [tid={tid}] 토지등기 링크 클릭 후 팝업 확인 중 오류: {e}")
    finally:
        dismiss_alert_if_present(driver, tid)
        try:
            if new_handle and new_handle in driver.window_handles:
                driver.switch_to.window(new_handle)
                driver.close()
        except Exception:
            pass
        try:
            if detail_tab_handle in driver.window_handles:
                driver.switch_to.window(detail_tab_handle)
            elif driver.window_handles:
                driver.switch_to.window(driver.window_handles[0])
        except Exception:
            pass
    return pdf_url


def download_pdf(session, pdf_url):
    try:
        resp = session.get(pdf_url, timeout=20)
        resp.raise_for_status()
        ctype = resp.headers.get("Content-Type", "").lower()
        if "pdf" in ctype or pdf_url.lower().endswith(".pdf"):
            return resp.content
        print(f"  PDF가 아닌 응답입니다: {pdf_url} (Content-Type={ctype})")
        return None
    except Exception as e:
        print(f"  PDF 다운로드 실패 ({pdf_url}): {e}")
        return None


def fetch_pdf_bytes_via_browser(driver, pdf_url, tid=None, timeout=25):
    driver.set_script_timeout(timeout)
    script = """
    var url = arguments[0];
    var callback = arguments[arguments.length - 1];
    fetch(url, {credentials: 'include'})
      .then(function(r) {
        var ctype = (r.headers.get('Content-Type') || '').toLowerCase();
        return r.arrayBuffer().then(function(buf) { return {ctype: ctype, buf: buf}; });
      })
      .then(function(result) {
        var bytes = new Uint8Array(result.buf);
        var binary = '';
        var chunk = 0x8000;
        for (var i = 0; i < bytes.length; i += chunk) {
          binary += String.fromCharCode.apply(null, bytes.subarray(i, i + chunk));
        }
        callback({ok: true, ctype: result.ctype, b64: btoa(binary)});
      })
      .catch(function(err) { callback({ok: false, error: String(err)}); });
    """
    try:
        result = driver.execute_async_script(script, pdf_url)
    except Exception as e:
        print(f"  [tid={tid}] 브라우저 fetch()로 PDF 요청 중 오류: {e}")
        return None

    if not result or not result.get("ok"):
        print(f"  [tid={tid}] 브라우저 fetch() 실패: {result.get('error') if result else '응답 없음'}")
        return None

    ctype = result.get("ctype", "")
    try:
        pdf_bytes = base64.b64decode(result["b64"])
    except Exception as e:
        print(f"  [tid={tid}] PDF base64 디코딩 실패: {e}")
        return None

    if not pdf_bytes.startswith(b"%PDF"):
        preview = pdf_bytes[:200].decode("utf-8", errors="replace").strip()
        print(f"  [tid={tid}] PDF가 아닌 응답입니다 (Content-Type={ctype}, %PDF로 시작하지 않음). "
              f"응답 앞부분: {preview!r}")
        return None
    return pdf_bytes


def extract_registry_summary_from_pdf(pdf_bytes):
    if not pdf_bytes or pdfplumber is None:
        return ""
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            if not pdf.pages:
                return ""
            page_texts = [(page.extract_text() or "") for page in pdf.pages]
    except Exception as e:
        print(f"  토지등기 PDF 파싱 오류: {e}")
        return ""

    summary_start_idx = None
    for i, t in enumerate(page_texts):
        if "주요 등기사항 요약" in t:
            summary_start_idx = i
            break

    if summary_start_idx is not None:
        full_text = "\n".join(page_texts[summary_start_idx:]).strip()
    else:
        full_text = page_texts[-1].strip()

    m = re.search(
        r"1\.\s*소유지분현황.*?(?=2\.\s*소유지분을\s*제외한|\Z)",
        full_text,
        re.DOTALL,
    )
    if m:
        return m.group(0).strip()
    return full_text


def extract_parcel_texts(soup):
    parcels = []
    rows_tbody = soup.select_one("table.ViewTbl tbody.rowsTbody")
    if not rows_tbody:
        return parcels
    for tr in rows_tbody.find_all("tr"):
        first_td = tr.find("td")
        if not first_td:
            continue
        first_div = first_td.find("div")
        if not first_div:
            continue
        raw = first_div.get_text(separator=" ", strip=True)
        parcel = raw.split("[")[0].strip()
        if parcel:
            parcels.append(parcel)
    return parcels


def build_full_parcel_address(base_address, parcel_text):
    if not base_address or not parcel_text:
        return parcel_text or base_address or ""
    primary = re.sub(r"\s*외\s*\d+\s*필지\s*$", "", base_address).strip()
    prefix = re.sub(r"\S*(동|리|가)\s*(산)?\d+(-\d+)?\s*$", "", primary).strip()
    if prefix:
        return f"{prefix} {parcel_text}".strip()
    return parcel_text


def extract_notes_from_detail(soup):
    parts = []
    for tr in soup.select("table.ViewTbl tr"):
        th = tr.find("th")
        td = tr.find("td")
        if not th or not td:
            continue
        label = th.get_text(separator=" ", strip=True)
        if ("현황" in label and "위치" in label) or label == "참고사항":
            for mobile_span in td.select("span.mobileTitle, span.show-mobile"):
                mobile_span.decompose()
            text = td.get_text(separator="\n", strip=True)
            if text:
                parts.append(f"[{label}]\n{text}")
    return "\n\n".join(parts)


def extract_sale_object_type(soup):
    for th in soup.find_all("th"):
        if th.get_text(strip=True) == "매각물건":
            td = th.find_next_sibling("td")
            if td:
                td_copy = BeautifulSoup(str(td), "html.parser")
                for mobile_span in td_copy.select("span.mobileTitle, span.show-mobile"):
                    mobile_span.decompose()
                return td_copy.get_text(strip=True)
    return ""


def extract_parcel_rows_pa(soup):
    target_table = None
    for table in soup.find_all("table"):
        header_cells = [th.get_text(strip=True) for th in table.find_all("th")]
        if "번호" in header_cells and "종별(지목)" in header_cells and "비고" in header_cells:
            target_table = table
            break
    if target_table is None:
        return []

    rows = []
    for tr in target_table.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) < 5:
            continue
        share_col_text = tds[3].get_text(separator=" ", strip=True)
        remark_text = tds[4].get_text(strip=True)
        if not share_col_text and not remark_text:
            continue
        rows.append((share_col_text, remark_text))
    return rows


def is_pa_share_parcel(share_col_text, remark_text=""):
    combined = f"{share_col_text or ''} {remark_text or ''}"
    return ("지분" in combined) or ("분의" in combined)


def extract_notes_from_detail_pa(soup):
    LABELS_TO_INCLUDE = ("이용현황", "기타")
    parts = []
    for tr in soup.select("table.ViewTbl tr"):
        th = tr.find("th")
        td = tr.find("td")
        if not th or not td:
            continue
        label = th.get_text(strip=True)
        if label in LABELS_TO_INCLUDE:
            text = td.get_text(separator="\n", strip=True)
            if text:
                parts.append(f"[{label}]\n{text}")
    return "\n\n".join(parts)


def extract_pa_pdf_url(soup):
    candidates = soup.select("a.btn-fileView[data-filelink]")
    link = None
    for c in candidates:
        if "토지등기" in c.get_text(strip=True):
            link = c
            break
    if link is None:
        for c in candidates:
            if "tp=I" in (c.get("data-filelink") or ""):
                link = c
                break
    if not link:
        return None
    filelink = (link.get("data-filelink") or "").strip()
    if not filelink:
        return None

    query_str = filelink.split("?", 1)[1] if "?" in filelink else ""
    qs = parse_qs(query_str)
    file_path_vals = qs.get("filePath")
    if file_path_vals and file_path_vals[0]:
        raw_path = unquote(file_path_vals[0])
        if not raw_path.startswith("/"):
            raw_path = "/" + raw_path
        return SITE_BASE + raw_path

    if filelink.startswith("http"):
        return filelink
    path = filelink if filelink.startswith("/") else "/pa/" + filelink
    return SITE_BASE + path


def build_pa_parcel_addresses(base_address, parcel_texts, category=None):
    addr = base_address or ""
    if category and addr.endswith(category):
        addr = addr[: -len(category)].strip()
    addr = re.sub(r"\s*외\s*\d+\s*필지\s*$", "", addr).strip()

    prefix = ""
    for p in parcel_texts:
        if p and addr.endswith(p):
            prefix = addr[: -len(p)].strip()
            break

    return [f"{prefix} {p}".strip() if prefix else p for p in parcel_texts]


def split_multi_parcel_address_pa(base_address, category=None):
    addr = (base_address or "").strip()
    if category and addr.endswith(category):
        addr = addr[: -len(category)].strip()
    if not addr:
        return []
    if "," not in addr:
        return [addr]

    parts = [p.strip() for p in addr.split(",") if p.strip()]
    if len(parts) < 2:
        return [addr]

    m = re.search(r"(산?\d+(?:-\d+)?)\s*$", parts[0])
    if not m:
        return [addr]

    prefix = parts[0][: m.start()].strip()
    first_parcel = m.group(1)
    result = [f"{prefix} {first_parcel}".strip() if prefix else first_parcel]
    for p in parts[1:]:
        if re.fullmatch(r"산?\d+(?:-\d+)?", p):
            result.append(f"{prefix} {p}".strip() if prefix else p)
        else:
            result.append(p)
    return result


def fetch_detail_info_pa(driver, tid, chk_no, tot_no):
    detail_url = f"{SITE_BASE}/pa/paView.php?cltrNo={tid}&chkNo={chk_no}&TotNo={tot_no}"
    parcel_texts = []
    notes_text = ""
    토지등기_요약 = ""
    지분여부 = ""
    original_handle = driver.current_window_handle
    detail_handle = None
    try:
        time.sleep(random.uniform(*DETAIL_PAGE_DELAY))
        detail_handle, html = open_detail_tab(driver, detail_url, tid=tid)
        if not detail_handle:
            return parcel_texts, notes_text, 토지등기_요약, 지분여부
        print(f"  [{TYPE_B_LABEL}][tid={tid}] 상세페이지(브라우저 탭) 로드 완료, 길이 {len(html)}자")
        soup = BeautifulSoup(html, "html.parser")

        parcel_rows = extract_parcel_rows_pa(soup)
        if parcel_rows:
            is_jibun = any(is_pa_share_parcel(share, remark) for share, remark in parcel_rows)
            parcel_texts = [
                remark for share, remark in parcel_rows
                if remark and not is_pa_share_parcel(share, remark)
            ]
        else:
            print(f"  [{TYPE_B_LABEL}][tid={tid}] 필지별 표를 찾지 못했습니다.")
            is_jibun = False
        지분여부 = "지분" if is_jibun else ""
        print(f"  [{TYPE_B_LABEL}][tid={tid}] 필지 {len(parcel_rows)}개, 지분여부={'지분' if is_jibun else '아님'}")

        notes_text = extract_notes_from_detail_pa(soup)

        if is_jibun:
            pdf_url = extract_pa_pdf_url(soup)
            if pdf_url:
                print(f"  [{TYPE_B_LABEL}][tid={tid}] 토지등기 PDF URL 확인: {pdf_url}")
                time.sleep(random.uniform(*PDF_DOWNLOAD_DELAY))
                pdf_bytes = fetch_pdf_bytes_via_browser(driver, pdf_url, tid=tid)
                if pdf_bytes:
                    print(f"  [{TYPE_B_LABEL}][tid={tid}] 토지등기 PDF 다운로드 완료 ({len(pdf_bytes)} bytes)")
                    토지등기_요약 = extract_registry_summary_from_pdf(pdf_bytes)
                    if not 토지등기_요약:
                        print(f"  [{TYPE_B_LABEL}][tid={tid}] 토지등기 PDF에서 텍스트를 추출하지 못했습니다.")
                else:
                    print(f"  [{TYPE_B_LABEL}][tid={tid}] 토지등기 PDF 다운로드 실패 (URL: {pdf_url})")
            else:
                print(f"  [{TYPE_B_LABEL}][tid={tid}] 지분 물건인데 토지등기 PDF 링크(data-filelink)를 찾지 못했습니다.")
    except Exception as e:
        print(f"[{TYPE_B_LABEL}][tid={tid}] 상세페이지 처리 중 오류: {e}")
    finally:
        close_detail_tab(driver, detail_handle, original_handle, tid=tid)

    return parcel_texts, notes_text, 토지등기_요약, 지분여부


def fetch_detail_info(driver, session, tid, chk_no, tot_no, base_address):
    detail_url = f"{SITE_BASE}/ca/caView.php?tid={tid}&chkNo={chk_no}&TotNo={tot_no}"
    parcel_addresses = []
    notes_text = ""
    토지등기_요약 = ""
    지분여부 = ""
    original_handle = driver.current_window_handle
    detail_handle = None
    try:
        time.sleep(random.uniform(*DETAIL_PAGE_DELAY))
        detail_handle, html = open_detail_tab(driver, detail_url, tid=tid)
        if not detail_handle:
            return parcel_addresses, notes_text, 토지등기_요약, 지분여부
        print(f"  [tid={tid}] 상세페이지(브라우저 탭) 로드 완료, 길이 {len(html)}자")
        soup = BeautifulSoup(html, "html.parser")

        parcel_texts = extract_parcel_texts(soup)
        if parcel_texts:
            parcel_addresses = [build_full_parcel_address(base_address, p) for p in parcel_texts]
        else:
            print(f"  [tid={tid}] '매각 물건 현황' 표에서 필지 정보를 찾지 못했습니다.")

        notes_text = extract_notes_from_detail(soup)
        if not notes_text:
            print(f"  [tid={tid}] '현황·위치·주변환경'/'참고사항' 텍스트를 찾지 못했습니다.")

        sale_object_type = extract_sale_object_type(soup)
        is_jibun = "지분" in sale_object_type
        지분여부 = "지분" if is_jibun else ""
        if sale_object_type:
            print(f"  [tid={tid}] 매각물건: {sale_object_type} -> 지분여부={'지분' if is_jibun else '아님'}")
        else:
            print(f"  [tid={tid}] 상세페이지에서 '매각물건' 항목을 찾지 못해 지분 여부를 판단하지 못했습니다.")

        if is_jibun:
            pdf_url = click_registry_pdf_link(driver, detail_handle, tid=tid)
            if pdf_url:
                time.sleep(random.uniform(*PDF_DOWNLOAD_DELAY))
                pdf_bytes = fetch_pdf_bytes_via_browser(driver, pdf_url, tid=tid)
                if pdf_bytes:
                    print(f"  [tid={tid}] 토지등기 PDF 다운로드 완료 ({len(pdf_bytes)} bytes)")
                    토지등기_요약 = extract_registry_summary_from_pdf(pdf_bytes)
                    if not 토지등기_요약:
                        print(f"  [tid={tid}] 토지등기 PDF 마지막 페이지에서 텍스트를 추출하지 못했습니다.")
                else:
                    print(f"  [tid={tid}] 토지등기 PDF 다운로드 실패 (URL: {pdf_url})")
            else:
                print(f"  [tid={tid}] 지분 물건인데 토지등기 PDF URL을 얻지 못했습니다.")

    except Exception as e:
        print(f"[tid={tid}] 상세페이지 처리 중 오류: {e}")
    finally:
        close_detail_tab(driver, detail_handle, original_handle, tid=tid)

    return parcel_addresses, notes_text, 토지등기_요약, 지분여부


def click_page_number(driver, target_page):
    buttons = driver.find_elements(By.CSS_SELECTOR, "#paging div.pageBtn")
    for b in buttons:
        if b.text.strip() == str(target_page):
            driver.execute_script("arguments[0].click();", b)
            return True
    return False


def click_next_block(driver):
    buttons = driver.find_elements(By.CSS_SELECTOR, "#paging div.pageBtn")
    for b in buttons:
        if b.text.strip() == "다음":
            driver.execute_script("arguments[0].click();", b)
            return True
    return False


def go_to_page(driver, target_page, prev_first_tid):
    moved = click_page_number(driver, target_page)
    if not moved:
        if not click_next_block(driver):
            return False
        time.sleep(random.uniform(*PAGE_CLICK_DELAY))
        moved = click_page_number(driver, target_page)
        if not moved:
            return False

    try:
        WebDriverWait(driver, 15).until(lambda d: get_first_tid(d) != prev_first_tid)
    except TimeoutException:
        print(f"경고: {target_page} 페이지로 이동 후 목록 변경을 확인하지 못했습니다.")
    return True


def main():
    collect_mode = prompt_collect_mode()
    apsl_amt_end_value = prompt_apsl_amt_end()
    login_id, login_pw = LOGIN_ID, LOGIN_PW

    existing_case_numbers = set()
    existing_rows = []

    # 예약(schedule) 실행인지 그 외(수동 실행 등)인지에 따라 완전히 다른 탭을 쓴다.
    # 예약 실행끼리는(type_a/type_b) 서로 같은 주의 같은 탭을 공유해서 누적되지만,
    # 사람이 직접 돌리는 실행은 그 탭을 절대 건드리지 않고 "실행일자" 전용 탭에
    # 따로 쌓는다 - 예약 실행이 아직 진행 중이거나 막 끝난 시점에 수동 실행을 해도,
    # 서로 다른 탭에 쓰기 때문에 한쪽이 다른 쪽을 덮어쓰는 사고 자체가 구조적으로
    # 발생할 수 없다(2026-08-28~29에 실제로 겪은, "나중에 끝난 실행이 시트 전체를
    # 자기 것으로 덮어써서 먼저 쌓인 데이터가 사라지는" 문제의 근본 해결책).
    is_scheduled_run = (TRIGGER_TYPE_ENV == "schedule")

    google_sh = None
    current_week_ws = None
    prev_week_label = None
    current_week_label = None
    if GOOGLE_SHEETS_ENABLED:
        오늘_date = date.today()
        if is_scheduled_run:
            current_week_label = _week_tab_name(오늘_date)
            prev_week_label = _week_tab_name(오늘_date - timedelta(days=7))
            print(f"\n=== 구글 시트 탭(예약 실행 전용): {current_week_label} ===")
            print(f"이 탭에 이미 있는 사건번호는 상세페이지/PDF 요청 없이 건너뜁니다. "
                  f"지난주 탭({prev_week_label})과 비교해 매각완료로 추정되는 물건 수를 실행이 끝날 때 안내합니다.")
        else:
            current_week_label = _manual_tab_name(오늘_date)
            prev_week_label = _week_tab_name(오늘_date - timedelta(days=7))
            print(f"\n=== 구글 시트 탭(수동 실행 전용): {current_week_label} ===")
            print(f"수동 실행이라 예약 전용 탭과는 별도의 탭에 저장됩니다 - 예약 실행 데이터에는 "
                  f"전혀 영향을 주지 않습니다. 같은 날 다시 수동 실행하면 이 탭에 이어서 누적됩니다.")
        print(f"새 물건 {AUTOSAVE_EVERY_N_ITEMS}건마다 자동으로 중간 저장합니다.")
        if os.path.exists(GOOGLE_CREDENTIALS_PATH):
            print(f"서비스 계정 키 파일: {GOOGLE_CREDENTIALS_PATH} (이 실행 환경에서 찾음)")
        else:
            print(f"경고: 서비스 계정 키 파일을 찾지 못했습니다. GOOGLE_SA_KEY_JSON 시크릿이 "
                  f"올바르게 설정되어 있는지 확인하세요 - 구글 시트 저장이 건너뛰어지고 로컬 xlsx만 저장됩니다.")
        google_sh = get_google_spreadsheet()
        if google_sh is not None:
            current_week_ws = get_or_create_tab(google_sh, current_week_label)
            if current_week_ws is not None:
                existing_case_numbers, existing_rows = load_existing_data_from_gsheet(current_week_ws)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "data"
    sheet.append(TEMPLATE_HEADERS)
    if FETCH_DETAIL_DOCS:
        sheet.cell(row=1, column=AC_COL, value="토지등기_요약(지분물건만, 임시)")

    # 로컬 엑셀 파일명은 실제 사용된 탭 이름 그대로 ".xlsx"만 붙인다(탭 이름이
    # current_week_label 하나로 이미 예약/수동 여부까지 반영되어 있으므로).
    # GOOGLE_SHEETS_ENABLED=False라 current_week_label이 없을 때는 실행일자
    # 기준 이름으로 대체한다(하위 호환).
    파일명 = f"{current_week_label or _manual_tab_name(date.today())}.xlsx"
    save_path = os.path.join(SAVE_DIR, 파일명)
    if os.path.exists(save_path):
        base, ext = os.path.splitext(save_path)
        save_path = f"{base}_A{ext}"

    for row_values, ac_value in existing_rows:
        append_row_with_format(sheet, row_values, ac_value)
    if existing_rows:
        기존_출처 = f"이번 주 구글 시트 탭({current_week_label})" if GOOGLE_SHEETS_ENABLED else "이전 파일"
        print(f"{기존_출처}에서 기존 물건 {len(existing_rows)}건을 가져왔습니다 "
              f"(사건번호 {len(existing_case_numbers)}종). 이미 있는 사건번호는 상세페이지/PDF 요청 없이 건너뜁니다.")

    gsheet_sync_state = {"synced_rows": 1 + len(existing_rows)}

    print(f"\n검색 조건: 물건종류코드={CTGR_VALUE}, 매각기일={start_date_str}~{end_date_str}, "
          f"감정가상한={int(apsl_amt_end_value):,}원")
    if FETCH_DETAIL_DOCS and pdfplumber is None:
        print("경고: pdfplumber 모듈이 설치되어 있지 않아 PDF 텍스트 추출이 동작하지 않습니다.")
    print()

    driver = get_driver()
    seen_tids = set()
    doc_session = None
    total_new_count = 0
    total_skip_existing = 0

    scanned_case_numbers_ca = set()
    scanned_case_numbers_pa = set()
    ca_scan_complete = False
    pa_scan_complete = False

    try:
        driver.get(BASE_URL)
        wait_for_list(driver)
        time.sleep(1)

        if FETCH_DETAIL_DOCS:
            login(driver, login_id, login_pw)
            time.sleep(1)

        if collect_mode in ("type_a", "both"):
            share_tids_ca = set()
            if FETCH_DETAIL_DOCS:
                print("\n[사전조회] '지분입찰 물건' 필터로 지분 물건 Tid만 먼저 수집합니다 (팝업 없음)...")
                share_tids_ca = collect_share_tids(driver, apsl_amt_end_value)

            apply_filters(driver, apsl_amt_end_value)
            wait_for_list(driver)
            set_page_size(driver)
            wait_for_list(driver)

            if FETCH_DETAIL_DOCS:
                doc_session = build_requests_session(driver)

            page_num = 1
            while page_num <= MAX_PAGES:
                print(f"현재 {page_num} 페이지 데이터 추출 중...")
                wait_for_list(driver)
                time.sleep(0.5)
                rows = get_rows(driver)

                if not rows:
                    print("검색 결과가 없습니다. 종료합니다.")
                    break

                prev_first_tid = rows[0].get_attribute("data-tid")
                tot_no = len(rows)
                page_extracted_count = 0
                for chk_no, row in enumerate(rows, start=1):
                    tid = None
                    try:
                        tid = row.get_attribute("data-tid")
                        if tid in seen_tids:
                            continue
                        seen_tids.add(tid)
                        data = extract_row(row, tid, chk_no=chk_no, tot_no=tot_no)
                        if not data:
                            continue
                        page_extracted_count += 1

                        사건번호_key = (data[1] or "").strip()
                        if 사건번호_key:
                            scanned_case_numbers_ca.add(사건번호_key)
                        if 사건번호_key and 사건번호_key in existing_case_numbers:
                            total_skip_existing += 1
                            continue

                        base_address = data[2]

                        needs_detail = (tid in share_tids_ca) or is_multi_parcel(base_address)

                        parcel_addrs, 등기요약 = [], ""
                        if FETCH_DETAIL_DOCS and needs_detail and doc_session is not None:
                            parcel_addrs, notes_text, 등기요약, 지분여부_상세 = fetch_detail_info(
                                driver, doc_session, tid, chk_no, tot_no, base_address
                            )
                            if 지분여부_상세:
                                data[13] = 지분여부_상세
                            elif tid in share_tids_ca:
                                data[13] = "지분"
                            if notes_text:
                                if data[3] and data[3] != "-":
                                    data[3] = f"{data[3]}\n\n{notes_text}"
                                else:
                                    data[3] = notes_text
                        else:
                            if tid in share_tids_ca:
                                data[13] = "지분"
                            data[14] = base_address

                        total_new_count += 1
                        if len(parcel_addrs) >= 2:
                            for addr in parcel_addrs:
                                row_data = list(data)
                                row_data[14] = addr
                                print(row_data)
                                append_row_with_format(sheet, row_data, 등기요약)
                        else:
                            if parcel_addrs:
                                data[14] = parcel_addrs[0]
                            print(data)
                            append_row_with_format(sheet, data, 등기요약)

                        maybe_autosave(sheet, current_week_ws, save_path, total_new_count, gsheet_sync_state)
                    except StaleElementReferenceException:
                        raise
                    except Exception as e:
                        print(f"[tid={tid}] 이 물건 처리 중 오류가 발생해 건너뜁니다: {e}")
                        continue

                if page_extracted_count == 0 and page_num > 1:
                    print("이 페이지에서 추출된 데이터가 없어 종료합니다.")
                    break

                next_page = page_num + 1
                if not go_to_page(driver, next_page, prev_first_tid):
                    print("마지막 페이지에 도달했습니다.")
                    break
                page_num = next_page

            ca_scan_complete = True

        if collect_mode in ("type_b", "both"):
            if collect_mode == "both":
                print(f"\n{TYPE_A_LABEL} 물건 수집 완료 (총 {sheet.max_row - 1}행). {TYPE_B_LABEL} 물건 수집을 시작합니다...")
            else:
                print(f"\n{TYPE_B_LABEL} 물건 수집을 시작합니다...")
            seen_tids_pa = set()
            pa_list_ok = True
            try:
                driver.get(PA_BASE_URL)
                wait_for_list(driver)
                time.sleep(1)
            except Exception as e:
                print(f"{TYPE_B_LABEL} 목록 페이지 접속 실패: {e} - {TYPE_B_LABEL} 수집을 건너뜁니다.")
                pa_list_ok = False

            pa_all_categories_ok = True
            if pa_list_ok:
                for prptdvsn_value, prptdvsn_label in PA_PRPTDVSN_VALUES:
                    try:
                        print(f"\n[{TYPE_B_LABEL}] 자산구분 '{prptdvsn_label}' 검색 시작...")
                        apply_filters_pa(driver, apsl_amt_end_value, prptdvsn_value)
                        wait_for_list(driver)
                        set_page_size(driver)
                        wait_for_list(driver)

                        pa_page_num = 1
                        while pa_page_num <= MAX_PAGES:
                            print(f"[{TYPE_B_LABEL}/{prptdvsn_label}] 현재 {pa_page_num} 페이지 데이터 추출 중...")
                            wait_for_list(driver)
                            time.sleep(0.5)
                            rows = get_rows(driver)

                            if not rows:
                                print(f"[{TYPE_B_LABEL}/{prptdvsn_label}] 검색 결과가 없습니다.")
                                break

                            prev_first_tid_pa = rows[0].get_attribute("data-tid")
                            tot_no = len(rows)
                            page_extracted_count = 0
                            for chk_no, row in enumerate(rows, start=1):
                                tid = None
                                try:
                                    tid = row.get_attribute("data-tid")
                                    if tid in seen_tids_pa:
                                        continue
                                    seen_tids_pa.add(tid)
                                    data = extract_row_pa(row, tid, chk_no=chk_no, tot_no=tot_no)
                                    if not data:
                                        continue
                                    page_extracted_count += 1

                                    사건번호_key = (data[1] or "").strip()
                                    if 사건번호_key:
                                        scanned_case_numbers_pa.add(사건번호_key)
                                    if 사건번호_key and 사건번호_key in existing_case_numbers:
                                        total_skip_existing += 1
                                        continue

                                    base_address = data[2]
                                    needs_detail_pa = (data[13] == "지분")

                                    parcel_addrs = split_multi_parcel_address_pa(
                                        base_address, category=data[4]
                                    )
                                    등기요약 = ""

                                    if FETCH_DETAIL_DOCS and needs_detail_pa:
                                        parcel_texts_raw, notes_text, 등기요약, 지분여부_상세 = \
                                            fetch_detail_info_pa(driver, tid, chk_no, tot_no)
                                        table_addrs = build_pa_parcel_addresses(
                                            base_address, parcel_texts_raw, category=data[4]
                                        ) if parcel_texts_raw else []
                                        if len(table_addrs) >= 2:
                                            parcel_addrs = table_addrs
                                        elif table_addrs:
                                            parcel_addrs = table_addrs
                                        if 지분여부_상세:
                                            data[13] = 지분여부_상세
                                        if notes_text:
                                            if data[3] and data[3] != "-":
                                                data[3] = f"{data[3]}\n\n{notes_text}"
                                            else:
                                                data[3] = notes_text

                                    total_new_count += 1
                                    if len(parcel_addrs) >= 2:
                                        for addr in parcel_addrs:
                                            row_data = list(data)
                                            row_data[14] = addr
                                            print(row_data)
                                            append_row_with_format(sheet, row_data, 등기요약)
                                    else:
                                        if parcel_addrs:
                                            data[14] = parcel_addrs[0]
                                        print(data)
                                        append_row_with_format(sheet, data, 등기요약)

                                    maybe_autosave(sheet, current_week_ws, save_path, total_new_count, gsheet_sync_state)
                                except StaleElementReferenceException:
                                    raise
                                except Exception as e:
                                    print(f"[{TYPE_B_LABEL}/{prptdvsn_label}][tid={tid}] 이 물건 처리 중 오류가 발생해 건너뜁니다: {e}")
                                    continue

                            if page_extracted_count == 0 and pa_page_num > 1:
                                print(f"[{TYPE_B_LABEL}/{prptdvsn_label}] 이 페이지에서 추출된 데이터가 없어 이 카테고리 수집을 종료합니다.")
                                break

                            next_page_pa = pa_page_num + 1
                            if not go_to_page(driver, next_page_pa, prev_first_tid_pa):
                                print(f"[{TYPE_B_LABEL}/{prptdvsn_label}] 마지막 페이지에 도달했습니다.")
                                break
                            pa_page_num = next_page_pa

                    except Exception as e:
                        print(f"[{TYPE_B_LABEL}/{prptdvsn_label}] 이 카테고리 처리 중 오류가 발생해 건너뜁니다: {e}")
                        pa_all_categories_ok = False
                        continue

                print(f"\n{TYPE_B_LABEL} 물건 수집 완료 (전체 저장된 행 {sheet.max_row - 1}행).")

                if pa_all_categories_ok:
                    pa_scan_complete = True

    except KeyboardInterrupt:
        print("사용자가 실행을 중지했습니다.")
    except StaleElementReferenceException:
        print("페이지 갱신 중 요소 참조가 끊겼습니다. 지금까지 수집된 데이터로 저장합니다.")
    finally:
        if GOOGLE_SHEETS_ENABLED and current_week_ws is not None:
            sync_to_google_sheet(sheet, current_week_ws)

            if is_scheduled_run and (ca_scan_complete or pa_scan_complete) and google_sh is not None:
                prev_week_ws = get_tab_if_exists(google_sh, prev_week_label)
                prev_ca_set, prev_pa_set = load_case_numbers_by_type_from_gsheet(prev_week_ws)
                sold_out_ca = (prev_ca_set - scanned_case_numbers_ca) if ca_scan_complete else set()
                sold_out_pa = (prev_pa_set - scanned_case_numbers_pa) if pa_scan_complete else set()
                sold_out_total = len(sold_out_ca) + len(sold_out_pa)
                if sold_out_total:
                    print(f"[참고] 지난주 탭({prev_week_label})에는 있었는데 이번 주 스캔 결과에는 보이지 않는 "
                          f"물건 {sold_out_total}건(매각완료 등으로 추정, {TYPE_A_LABEL} {len(sold_out_ca)}건 / "
                          f"{TYPE_B_LABEL} {len(sold_out_pa)}건)이 있습니다.")

            # 오래된 탭 자동 정리도 예약 실행일 때만 수행한다 - 수동 실행(수동 전용
            # 탭 이름 패턴이라 어차피 WEEK_TAB_PATTERN에 안 걸려 청소 대상이 되지도
            # 않지만) 도중에 매번 정리 API 호출까지 할 필요는 없으므로 함께 건너뛴다.
            if is_scheduled_run and google_sh is not None:
                cleanup_old_week_tabs(google_sh, GOOGLE_SHEET_WEEK_RETENTION)

        workbook.save(save_path)
        print(f"Excel 파일 저장 완료: {save_path} "
              f"(신규 {total_new_count}건 추가, 기존 {total_skip_existing}건 건너뜀, "
              f"전체 저장된 행 {sheet.max_row - 1}건)")
        driver.quit()


if __name__ == "__main__":
    main()
